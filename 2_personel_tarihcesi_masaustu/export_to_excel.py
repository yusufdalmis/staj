import sqlite3
import pandas as pd
import os

DB_PATH = r"d:\stajv2\2024_03\personel_veritabani.sqlite"
EXCEL_PATH = r"d:\stajv2\2024_03\Personel_Raporu.xlsx"

def export_data(db_path, excel_path, callback=None):
    if not os.path.exists(db_path):
        if callback: callback("Veritabanı bulunamadı.")
        return

    conn = sqlite3.connect(db_path)
    df_personel = pd.read_sql_query('''
        SELECT G.id, G.dosya_adi, G.yil, G.ay, P.isim || ' ' || P.soyisim AS ad_soyad, G.unvan, G.birim 
        FROM Gorev G 
        JOIN Personel P ON G.personel_id = P.id 
        ORDER BY G.yil, G.ay
    ''', conn)
    
    if df_personel.empty:
        if callback: callback("Kayıt yok.")
        conn.close()
        return

    # Komisyonları normal birimlerden ayıralım
    komisyon_keywords = ['KOMİSYONU', 'SORUMLULARI']
    mask = df_personel['birim'].str.contains('|'.join(komisyon_keywords), case=False, na=False)
    
    df_komisyon = df_personel[mask].copy()
    df_personel = df_personel[~mask].copy()

    df_personel['donem'] = df_personel['yil'].astype(str) + "_" + df_personel['ay'].astype(str).str.zfill(2)
    df_pivot = df_personel.drop_duplicates(subset=['ad_soyad', 'donem'], keep='last')
    
    # 1. PERSONEL TARIHCESI (Gerçek Veri, ffill yok)
    pivot_table = df_pivot.pivot(index='ad_soyad', columns='donem', values='birim')

    # 2. BIRIM OZET
    df_birimler = df_personel.groupby('birim').agg(
        ilk_kayit=('donem', 'min'),
        son_kayit=('donem', 'max')
    ).reset_index()

    def find_baskan(birim_adi):
        baskan_df = df_personel[(df_personel['birim'] == birim_adi) & (df_personel['unvan'].str.contains('Başkan|Koordinatör|Genel Sekreter', case=False, na=False))]
        if not baskan_df.empty:
            return baskan_df.iloc[-1]['ad_soyad']
        return "Bilinmiyor"
        
    df_birimler['guncel_yonetici'] = df_birimler['birim'].apply(find_baskan)
    df_birimler.rename(columns={
        'birim': 'Birim Adı',
        'ilk_kayit': 'Kuruluş / İlk Kayıt',
        'son_kayit': 'Son Kayıt',
        'guncel_yonetici': 'Güncel/Son Yönetici'
    }, inplace=True)

    # DİĞER ANALİZLER (Artık ffill olmadan, df_personel üzerinden)
    df_birim_kisi = df_personel.groupby(['birim', 'donem']).size().unstack(fill_value=0)
    df_toplam_personel = df_personel.groupby('donem').size().reset_index(name='Toplam Personel')
    df_unvan_dagilimi = df_personel.groupby(['unvan', 'donem']).size().unstack(fill_value=0)

    # EXCEL YAZDIRMA
    if callback: callback("Excel dosyası oluşturuluyor...")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_personel.drop(columns=['donem', 'id']).to_excel(writer, sheet_name='Tüm Kayıtlar (Ham)', index=False)
        df_birimler.to_excel(writer, sheet_name='Birimler Özet', index=False)
        
        # Yeni İstenen Sayfa 2: Birim Başkanları Tarihçesi
        baskan_df = df_personel[df_personel['unvan'].str.contains('Başkan|Koordinatör|Genel Sekreter', case=False, na=False)]
        if not baskan_df.empty:
            df_baskan_tarihce = baskan_df.pivot_table(index='birim', columns='donem', values='ad_soyad', aggfunc=lambda x: ' / '.join(x.dropna().unique()))
            df_baskan_tarihce.to_excel(writer, sheet_name='Birim Başkanları Tarihçesi')
            
        pivot_table.to_excel(writer, sheet_name='Personel Tarihçesi')
        
        # Yeni İstenen Sayfa 1: Personel Kariyer Yolu (Birim Geçişleri)
        career_paths = []
        for person, group in df_personel.sort_values(['yil', 'ay']).groupby('ad_soyad'):
            path = []
            for b in group['birim']:
                if not path or path[-1] != b:
                    path.append(b)
            career_paths.append({'ad_soyad': person, 'path': path})
            
        if career_paths:
            max_len = max(len(p['path']) for p in career_paths)
            records = []
            for p in career_paths:
                row = {'Personel Adı': p['ad_soyad']}
                for i in range(max_len):
                    row[f'Birim {i+1}'] = p['path'][i] if i < len(p['path']) else None
                records.append(row)
            df_kariyer = pd.DataFrame(records)
            df_kariyer.to_excel(writer, sheet_name='Kariyer Yolu', index=False)
        
        df_birim_kisi.to_excel(writer, sheet_name='Birim Başına Personel')
        df_toplam_personel.to_excel(writer, sheet_name='Toplam Personel Sayısı', index=False)
        df_unvan_dagilimi.to_excel(writer, sheet_name='Unvan Dağılımı')
        
        # Komisyonlar sayfası
        if not df_komisyon.empty:
            df_komisyon['donem'] = df_komisyon['yil'].astype(str) + "_" + df_komisyon['ay'].astype(str).str.zfill(2)
            df_komisyon.drop(columns=['id']).sort_values(['donem', 'birim']).to_excel(writer, sheet_name='Tüm Komisyonlar (Tarihçe)', index=False)
            
            # Güncel Komisyonlar Sayfası
            en_son_donem = df_komisyon['donem'].max()
            df_guncel_komisyon = df_komisyon[df_komisyon['donem'] == en_son_donem].drop(columns=['id', 'donem', 'yil', 'ay'])
            df_guncel_komisyon.sort_values('birim').to_excel(writer, sheet_name='Güncel Komisyonlar', index=False)
            
    # Renklendirme islemi (Açılıp tekrar kaydedilerek)
    if callback: callback("Renklendirme işlemi yapılıyor...")
    import openpyxl
    import hashlib
    from openpyxl.styles import PatternFill
    
    wb = openpyxl.load_workbook(excel_path)
    known_units = set(df_personel['birim'].unique()).union(set(df_komisyon['birim'].unique()))
    
    colors = [
        "FFB3BA", "FFDFBA", "FFFFBA", "BAFFC9", "BAE1FF", 
        "E8BAFF", "FFBAE1", "E2F0CB", "FFB7B2", "FFDAC1",
        "E0BBE4", "957DAD", "D291BC", "FEC8D8", "FFDFD3",
        "B5EAD7", "C7CEEA", "F1CBFF", "C8E6C9", "FFF9C4",
        "F4C2C2", "B0E0E6", "F5DEB3", "D8BFD8", "FFDEAD",
        "AFEEEE", "E6E6FA", "F08080", "E0FFFF", "FFB6C1",
        "98FB98", "FFDAB9", "FFE4B5", "F5F5DC", "F0E68C"
    ]
    
    def get_color(birim):
        hash_val = int(hashlib.md5(birim.encode('utf-8')).hexdigest(), 16)
        return colors[hash_val % len(colors)]
        
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for unit in known_units:
                        if cell.value == unit:
                            color_hex = get_color(unit)
                            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                            break

    wb.save(excel_path)
            
    msg = f"Excel raporu başarıyla oluşturuldu: {excel_path}"
    if callback: callback(msg)
    else: print(msg)
    
    conn.close()

if __name__ == "__main__":
    print("Excel'e aktarim baslatiliyor...")
    export_data(DB_PATH, EXCEL_PATH)
