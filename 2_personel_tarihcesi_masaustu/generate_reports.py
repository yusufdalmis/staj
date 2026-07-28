import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

DB_PATH = r"d:\stajv2\2024_03\personel_veritabani.sqlite"
OUTPUT_PATH = r"d:\stajv2\2024_03\Aylik_Cikti_Raporlari.xlsx"

def generate_monthly_reports(db_path, output_path, callback=None):
    if not os.path.exists(db_path):
        if callback: callback("Veritabanı bulunamadı!")
        return

    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query('''
        SELECT G.id, G.dosya_adi, G.yil, G.ay, P.isim || ' ' || P.soyisim AS ad_soyad, G.unvan, G.birim 
        FROM Gorev G 
        JOIN Personel P ON G.personel_id = P.id 
        ORDER BY G.yil, G.ay
    ''', conn)
    conn.close()

    if df.empty:
        if callback: callback("Kayıt yok.")
        return

    df['donem'] = df['yil'].astype(str) + "_" + df['ay'].astype(str).str.zfill(2)
    
    wb = Workbook()
    wb.remove(wb.active) # Varsayılan sayfayı sil
    
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Gri arka plan
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    donemler = df['donem'].unique()
    
    for donem in donemler:
        if callback: callback(f"Aylık Rapor Oluşturuluyor: {donem}")
        df_donem = df[df['donem'] == donem]
        
        ws = wb.create_sheet(title=str(donem))
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        
        ws.append(["ADI SOYADI", "UNVANI"])
        
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
        current_row = 2
        current_birim = None
        
        for _, row in df_donem.iterrows():
            if row['birim'] != current_birim:
                # PDF'teki sıraya göre yeni birim başlığı eklendi
                current_birim = row['birim']
                ws.append([current_birim, ""])
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                
                cell = ws.cell(row=current_row, column=1)
                cell.font = Font(bold=True)
                
                komisyon_keywords = ['KOMİSYON', 'KOMİTE', 'KURUL', 'SORUMLULARI']
                is_komisyon = any(k in current_birim.upper() for k in komisyon_keywords)
                
                if is_komisyon:
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Turuncu/Sarımsı
                else:
                    import hashlib
                    colors = [
                        "FFB3BA", "FFDFBA", "FFFFBA", "BAFFC9", "BAE1FF", 
                        "E8BAFF", "FFBAE1", "E2F0CB", "FFB7B2", "FFDAC1",
                        "E0BBE4", "957DAD", "D291BC", "FEC8D8", "FFDFD3",
                        "B5EAD7", "C7CEEA", "F1CBFF", "C8E6C9", "FFF9C4",
                        "F4C2C2", "B0E0E6", "F5DEB3", "D8BFD8", "FFDEAD",
                        "AFEEEE", "E6E6FA", "F08080", "E0FFFF", "FFB6C1",
                        "98FB98", "FFDAB9", "FFE4B5", "F5F5DC", "F0E68C"
                    ]
                    hash_val = int(hashlib.md5(current_birim.encode('utf-8')).hexdigest(), 16)
                    color = colors[hash_val % len(colors)]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                cell.border = thin_border
                ws.cell(row=current_row, column=2).border = thin_border
                current_row += 1
                
            # Kişiyi ekle
            ws.append([row['ad_soyad'], row['unvan']])
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=2).border = thin_border
            current_row += 1

    wb.save(output_path)
    msg = f"Baskıya hazır aylık raporlar başarıyla oluşturuldu: {output_path}"
    if callback: callback(msg)
    else:
        try: print(msg)
        except: print("Raporlar uretildi.")

if __name__ == "__main__":
    generate_monthly_reports(DB_PATH, OUTPUT_PATH)
